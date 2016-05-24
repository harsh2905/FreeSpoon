#!/usr/bin/python
# -*- coding: utf-8 -*-

import os, re
import MySQLdb
from datetime import datetime
from openpyxl import load_workbook

import sys
reload(sys)
sys.setdefaultencoding('utf-8')

import pdb

class Report():

	def __init__(self):
		self.conn = MySQLdb.connect(
			host='localhost', 
			user='root', 
			passwd='123456', 
			port=3306,
			charset='utf8')
		self.conn.select_db('FreeSpoon')
		self.cellDatas = {}

	def generate(self, path):
		if not os.path.exists(path):
			return
		wb = load_workbook(path)
		if wb is None:
			return
		ws = wb.active
		wss = wb.create_sheet()
		position = 0
		for r in range(1, ws.max_row + 1):
			repeatNum = 0
			for c in range(0, ws.max_column):
				cellName = chr(ord('A') + c) + str(r)
				newCellName = chr(ord('A') + c) + str(r + position)
				val = ws[cellName].value
				val = val if val is not None else ''
				val = str(val)
				if val is None or len(val.strip()) == 0:
					continue
				(newVal, repeatNum_) = self.calcExp(cellName, val)#TODO
				repeatNum = repeatNum_ if repeatNum_ > repeatNum else repeatNum
				self.copyStyles(ws[newCellName], wss[newCellName])
				wss[newCellName] = newVal
			if repeatNum > 0:
				position = position + repeatNum
				for _ in range(0, repeatNum):
					for c in range(0, ws.max_column):
						rawCellName = chr(ord('A') + c) + str(r)
						cellName_ = chr(ord('A') + c) + str(r + 1 + _)
						(newVal_, __) = self.calcExp(cellName_, rawCellName=rawCellName)
						self.copyStyles(ws[rawCellName], wss[cellName_])
						wss[cellName_] = newVal_
		for c in range(0, ws.max_column):
			colName = chr(ord('A') + c)
			rawWidth = ws.column_dimensions[colName].width
			if rawWidth is not None:
				wss.column_dimensions[colName].width = rawWidth
		for r in range(1, ws.max_row + 1):
			rawHeight = ws.row_dimensions[r].height
			if rawHeight is not None:
				wss.row_dimensions[r].height = rawHeight
		pdb.set_trace()
		wss.sheet_view.showGridLines = ws.sheet_view.showGridLines
		wb.save(path + '.' + datetime.strftime(datetime.now(), '%Y%m%d%H%M%S'))
		self.cellDatas = {}

	def copyStyles(self, cell, newCell):
		newCell.font = cell.font
		newCell.border = cell.border
		newCell.fill = cell.fill
		newCell.number_format = cell.number_format
		newCell.protection = cell.protection
		newCell.alignment = cell.alignment

	def calcExp(self, cellName, cellVal=None, rawCellName=None):
		# Database Column Name: SQL statement
		rexp = r'\{\{(.+)\}\}'
		if cellVal is None:
			if rawCellName is None:
				return ('', 0)
			cellData = self.cellDatas.get(rawCellName, None)
			if cellData is None:
				return ('', 0)
			sqlData = cellData.get('sqlData', None)
			template = cellData.get('template', None)
			columnName = cellData.get('columnName', None)
			rawRowIndex = int(rawCellName[1:])
			currentRowIndex = int(cellName[1:])
			lineNum = currentRowIndex - rawRowIndex
			if len(sqlData) <= lineNum:
				return (template, 0)
			rowData = sqlData[lineNum]
			val = rowData.get(columnName, None)
			val = val if val is not None else '' # !
			val = str(val)
			newVal = re.sub(rexp, val, template)
			return (newVal, 0)
		matchGroup = re.search(rexp, cellVal, re.M|re.I)
		if matchGroup is None:
			return (cellVal, 0)

		exp = matchGroup.group(1)

		args = exp.split(':')
		if len(args) <> 2:
			return (cellVal, 0)
		columnName = args[0]
		sqlStatement = args[1]
		sqlData = None
		if len(sqlStatement.strip()) == 0:
			columnIndex = ord(cellName[0])
			while (columnIndex > ord('A')):
				columnIndex = columnIndex - 1
				cellData = self.cellDatas.get(chr(columnIndex) + cellName[1:], None)
				if cellData is not None:
					sqlData = cellData.get('sqlData', None)
					break
		else:
			cur = self.conn.cursor(MySQLdb.cursors.DictCursor)
			cur.execute(sqlStatement)
			sqlData = cur.fetchall()
		if sqlData is None:
			return (cellVal, 0)
		cellData = {
			'sqlData': sqlData,
			'template': cellVal,
			'columnName': columnName
		}
		self.cellDatas[cellName] = cellData
		rowData = sqlData[0]
		val = rowData.get(columnName, None)
		val = val if val is not None else '' # !
		val = str(val)
		newVal = re.sub(rexp, val, cellVal)
		return (newVal, len(sqlData) - 1)


if __name__ == '__main__':
	r = Report()
	r.generate('../../1.xlsx')








